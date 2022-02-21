# -*- coding: utf-8 -*-


import os 
from pathlib import Path
from tkinter import E
import pandas as pd 
import xlrd
import glob
import openpyxl
from xls2xlsx import XLS2XLSX

def get_project_root():
    """Returns project root file path"""
    return str(Path(__file__).parent)

project_root=get_project_root()
print(project_root)
# 'JAN--22',give your folder name
Month_folder=os.path.join(project_root,"JAN-22")
# to get the month and year of the date
# a,b,month=Month_folder.split('\\')
X=[]
for subdir,dir,file in os.walk(Month_folder):
    for i in dir:
        X.append(os.path.join(Month_folder,i))

# All the excel

Excel_sheets=[]
for i in X:
  
 for a,b,c in os.walk(i):
    for n in c:
     Excel_sheets.append((os.path.join(i,n)))

# reading excel file 
df=pd.read_excel(Excel_sheets[3],skiprows=4)
# get the componet name 
for comp in df.columns:
    if 'Component' in comp:
        component=comp
# get the Month
for month in df.columns:
    if 'month' in month.lower():
        x,Month=month.split(':')
# reading the columns from 6th rows
df_new=pd.read_excel(Excel_sheets[3],skiprows=5)
# forward filling
df_new.ffill(inplace=True)
# getting only those columns which contains sum
df_new=df_new[df_new['Unnamed: 2']=='sum']
# for second format 
Datas={"Date":[]}
df_second_format=[]
# getting types of rejections 
rej_col=df_new.columns
for i in rej_col:
    if type(i)==int:
     Datas['Date'].append(i)
     key,value=component.split('-')
     Datas.setdefault(key,[])
     Datas[key].append(value)
for i in range(len(df_new)):
    for j in range(3,int(len(Datas['Date'])+3)):
     Datas.setdefault(df_new.iloc[i,1],[])
     Datas[df_new.iloc[i,1]].append(df_new.iloc[i,j])
final_df=pd.DataFrame(Datas)
final_df['Date']=final_df['Date'].apply(lambda x: str(x) +'-'+Month)
file_path=os.path.join(r'C:\Automation\JAN-22','converted.xlsx')
final_df.to_excel(file_path,index=False)


# start working on excel sheet
# must put data only as a true to get the values not the formula
# wb=openpyxl.load_workbook(Excel_sheets[1],data_only=True)
# # sheets=wb.get_sheet_names()
# # it will give the list of sheet names
# sheets=wb.sheetnames
# # we are interested in reading first sheet only so we will get working sheet by following way
# sheet=wb.get_sheet_by_name(sheets[0])
# print(sheet['K8'].value)
# # unfreeze the cells
# sheet.freeze_panes=None
# # unmerging cells
# for items in sorted(sheet.merged_cell_ranges):
#     print(items)
#     sheet.unmerge_cells(str(items) )

# # get highest rows and columns
# rows=sheet.max_row
# columns=sheet.max_column
# print(rows,columns)
# print(sheet['AI6'].value)

#  # get the componet id and and month

# for i in range(1,columns):
#   if 'Component' in  str(sheet.cell(row=5,column=i).value) :
#       component=str(sheet.cell(row=5,column=i).value)
# # get the month name 
# for i in range(1,columns):
#     if 'Month' in str(sheet.cell(row=5,column=i).value) :
#         Month=str(sheet.cell(row=5,column=i).value)
#         a,month=Month.split(':')

      

# Datas={}
# for j in range(4,columns):
#     # column j+1 is added to remove the last None value 
#     if sheet.cell(row=6,column=j+1).value !=None:
       
#        Datas.setdefault("Date",[])
#        Datas['Date'].append((sheet.cell(row=6,column=j).value))
#        # get the component IDs
#        key,value=component.split('-')
#        Datas.setdefault(key,[])
#        Datas[key].append(value)


# # data frame with two features
# df1=pd.DataFrame(Datas)
# for i in range(7,rows):
#     for j in range(4,4+len(df1)):
#         if sheet.cell(row=i,column=2).value==None:
#             pass
#         else:
#          Datas.setdefault(sheet.cell(row=i,column=2).value,[])# to get the key value
#          if sheet.cell(row=i,column=j).value==None:
#               Datas[sheet.cell(row=i,column=2).value].append('')
#          else:
#           Datas[sheet.cell(row=i,column=2).value].append(sheet.cell(row=i,column=j).value)
# #ws.Columns.EntireColumn.Hidden=False
# # sheet.Columns.EntireColumn.Hidden=False
# # wb.save('freezeExample.xlsx')
# df=pd.DataFrame(Datas)
# df['Date']=df['Date'].apply(lambda x: str(x) +'-'+month)
# file_path=os.path.join(r'C:\Automation\JAN-22','converted.xlsx')
# df.to_excel(file_path,index=False)



